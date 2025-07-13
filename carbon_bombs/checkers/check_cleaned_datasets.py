import itertools

import pandas as pd
from openpyxl import load_workbook

from carbon_bombs.conf import FPATH_OUT_ALL
from carbon_bombs.conf import FPATH_SRC_GEM_COAL
from carbon_bombs.conf import FPATH_SRC_GEM_GASOIL
from carbon_bombs.conf import FPATH_SRC_KHUNE_PAPER

# =====
# Load datasets to check values #

# Prepare khune paper Gasoil dataframe
cb_gasoil_source_df = pd.read_excel(
    FPATH_SRC_KHUNE_PAPER,
    sheet_name="Oil&Gas",
    skipfooter=4,
    skiprows=1,
)
cb_gasoil_source_df = cb_gasoil_source_df.loc[
    :, ["New", "Project", "Country", "Gt CO2"]
]
cb_gasoil_source_df.columns = [
    "New",
    "Project Name",
    "Country",
    "Potential emissions (GtCO2)",
]
cb_gasoil_source_df["Fuel"] = "Oil&Gas"
cb_gasoil_source_df["Country"] = cb_gasoil_source_df["Country"].replace(
    {
        "Russian Federation": "Russia",
        "Turkey": "Türkiye",
        "Saudi-Arabia": "Saudi Arabia",
        "Kuwait-Saudi-Arabia-Neutral Zone": "Kuwait",  # see Readme to get the details of this choice
    }
)

# Prepare khune paper Coal dataframe
cb_coal_source_df = pd.read_excel(
    FPATH_SRC_KHUNE_PAPER, sheet_name="Coal", skipfooter=3
)
cb_coal_source_df = cb_coal_source_df.loc[
    :, ["New", "Project Name", "Country", "Potential emissions (GtCO2)", "Fuel"]
]
cb_coal_source_df["Country"] = cb_coal_source_df["Country"].replace(
    {"Russian Federation": "Russia", "Turkey": "Türkiye"}
)

# merge both dataframe into one
cb_source_df = pd.concat([cb_coal_source_df, cb_gasoil_source_df])
cb_source_df["Project Name"] = cb_source_df["Project Name"].str.strip()
cb_source_df = cb_source_df.replace({"Türkiye": "Turkey"})


# GEM coal ans gasoil source
# gem_coal_df = pd.read_excel(FPATH_SRC_GEM_COAL, sheet_name="Global Coal Mine Tracker")
gem_coal_df = pd.read_excel(FPATH_SRC_GEM_COAL, sheet_name="GCMT Non-closed Mines")
gem_gasoil_df = pd.read_excel(
    FPATH_SRC_GEM_GASOIL, sheet_name="Main data", engine="openpyxl"
)


# ================== #
# Checkers functions #
# ================== #


def _check_size(cb_df):
    """Check CB df size"""
    if len(cb_df) == 425:
        return "✅ OK - carbon_bombs_info: length is 425\n"
    else:
        return (
            f"❌ KO - carbon_bombs_info: length different from 425 (n = {len(cb_df)})\n"
        )


def _check_all_cb_proj_available(cb_df, merge_df):
    """Check All projects are available"""
    if len(merge_df) == len(cb_df):
        return "✅ OK - carbon_bombs_info: all projects from Khune paper kept the same name\n"
    else:
        return "❌ KO - carbon_bombs_info: some projects did not kept the same name as the Khune paper names\n"


def _check_status(cb_df, merge_df):
    """Check status"""
    txt = ""

    if cb_df["Status_source_CB"].isna().sum() == 0:
        txt += "✅ OK - carbon_bombs_info: `Status_source_CB` has no missing value\n"
    else:
        txt += (
            f"❌ KO - carbon_bombs_info: `Status_source_CB` has some missing values\n"
        )

    if set(cb_df["Status_source_CB"].unique()) == {"operating", "not started"}:
        txt += "✅ OK - carbon_bombs_info: `Status_source_CB` has the right values {'operating', 'not started'}\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: `Status_source_CB` does not match the wanted values (found values: set(cb_df['Status_source_CB'].unique()))\n"

    if all((merge_df["New"] == "*") == (merge_df["Status_source_CB"] == "not started")):
        txt += "✅ OK - carbon_bombs_info: `Status_source_CB` kept the new project\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: `Status_source_CB` didnt keep the new projects as such\n"

    if all((merge_df["New"] != "*") == (merge_df["Status_source_CB"] == "operating")):
        txt += (
            "✅ OK - carbon_bombs_info: `Status_source_CB` kept the operating project\n"
        )
    else:
        txt += f"❌ KO - carbon_bombs_info: `Status_source_CB` didnt keep the operating projects as such\n"

    return txt


def _check_cb_names(cb_df):
    """Check CB names"""
    if set(cb_df["Project_name"]) == set(cb_source_df["Project Name"]):
        return "✅ OK - carbon_bombs_info: All Carbon bombs names were found \n"
    else:
        diff = set(cb_df["Project_name"]) - set(cb_source_df["Project Name"])
        return f"❌ KO - carbon_bombs_info: Some Carbon bombs are missing ({diff})\n"


def _check_cb_country(merge_df):
    """Check CB country"""
    if all((merge_df["Country"] == merge_df["Country"])):
        return (
            "✅ OK - carbon_bombs_info: All Carbon bombs countries were keep as such\n"
        )
    else:
        return f"❌ KO - carbon_bombs_info: Some Carbon bombs countries changed during the process\n"


def _check_cb_emissions(merge_df):
    """Check emissions"""
    if all(
        (
            merge_df["Potential emissions (GtCO2)"]
            == merge_df["Potential_GtCO2_source_CB"]
        )
    ):
        return "✅ OK - carbon_bombs_info: All Carbon bombs Potential emissions were keep as such \n"
    else:
        return f"❌ KO - carbon_bombs_info: Some Carbon bombs Potential emissions changed during the process\n"


def _check_cb_fuel(merge_df):
    """Check Fuel"""
    if all((merge_df["Fuel"] == merge_df["Fuel_type_source_CB"])):
        return (
            "✅ OK - carbon_bombs_info: All Carbon bombs Fuel type were keep as such \n"
        )
    else:
        return f"❌ KO - carbon_bombs_info: Some Carbon bombs Fuel type changed during the process\n"


def _check_units_mines_found_in_gem(units):
    """Check mines and units are all in cleaned df"""
    diff = (
        units
        # - set(gem_coal_df["Mine IDs"])
        - set(gem_coal_df["GEM Mine ID"])
        - set(gem_gasoil_df["Unit ID"])
        - {"No informations available on GEM"}
    )
    if len(diff) == 0:
        return "✅ OK - carbon_bombs_info: All mines and units were found\n"
    else:
        return f"❌ KO - carbon_bombs_info: Some mines or units are missing ({diff})\n"


def _check_gem_url_and_source_found_in_cb(url, cb_df):
    """Check gem wiki url"""
    diff = (
        url
        # - set(gem_coal_df["GEM Wiki Page (ENG)"])
        - set(gem_coal_df["GEM Wiki URLs"])
        - set(gem_gasoil_df["Wiki URL"])
        - {"No informations available on GEM"}
    )
    txt = ""

    if len(diff) == 0:
        txt += "✅ OK - carbon_bombs_info: All mines and units Wiki URL were found\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: Some mines or units Wiki URL are missing ({diff})\n"

    if all(
        (cb_df["GEM_url_source_GEM"] == "No informations available on GEM")
        == (cb_df["GEM_id_source_GEM"] == "No informations available on GEM")
    ):
        txt += "✅ OK - carbon_bombs_info: `GEM_url_source_GEM` and `GEM_id_source_GEM` missing information are for the same projects\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: `GEM_url_source_GEM` and `GEM_id_source_GEM` missing information are not for the same projects\n"

    return txt


def _check_cb_lat_long(cb_df):
    """Check CB lat and long"""
    txt = ""

    if cb_df["Latitude"].isna().sum() == 0:
        txt += "✅ OK - carbon_bombs_info: `Latitude` no missing values\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: `Latitude` missing values found\n"

    if cb_df["Longitude"].isna().sum() == 0:
        txt += "✅ OK - carbon_bombs_info: `Longitude` no missing values\n"
    else:
        txt += f"❌ KO - carbon_bombs_info: `Longitude` missing values found\n"

    if cb_df[["Latitude", "Longitude"]].duplicated().sum() == 0:
        txt += "✅ OK - carbon_bombs_info: No dupplicated values for Latitude, Longitude pair\n"
    else:
        txt += "❌ OK - carbon_bombs_info: there are some dupplicated values for Latitude, Longitude pairs\n"

    if cb_df["Latitude_longitude_source"].isna().sum() == 0:
        txt += "✅ OK - carbon_bombs_info: `Latitude_longitude_source` no missing values \n"
    else:
        txt += "❌ KO - carbon_bombs_info: `Latitude_longitude_source` missing values found\n"

    if set(cb_df["Latitude_longitude_source"]) == {"GEM", "Country CB", "Manual"}:
        txt += "✅ OK - carbon_bombs_info: `Latitude_longitude_source` values are {'GEM', 'Country CB', 'Manual'}\n"
    else:
        val = set(cb_df["Latitude_longitude_source"])
        txt += f"❌ KO - carbon_bombs_info: `Latitude_longitude_source` values are not the wanted ones (found {val})\n"

    return txt


def _check_cb_operators(cb_df):
    """Check CB Operators"""

    # - Operators_source_GEM :
    #     - Split |
    #     - merge GEM id and check same operators
    #     - show amount of "" or None
    # a = gem_coal_df[["Mine IDs", "Operators"]]
    a = gem_coal_df[["GEM Mine ID", "Operators"]]
    b = gem_gasoil_df[["Unit ID", "Operator"]]
    a.columns = ["ID", "Operator"]
    b.columns = ["ID", "Operator"]
    gem_id_df = pd.concat([a, b])

    def match_gem_id_to_operator(x):
        if x == ["No informations available on GEM"]:
            return ["None"]

        res = []
        for unit_id in x:
            ope = (
                gem_id_df.loc[gem_id_df["ID"] == unit_id, "Operator"]
                .fillna("None")
                .values[0]
            )
            res.append(ope)

        return res

    operators = cb_df["Operators_source_GEM"].fillna("None").str.split("|")
    diff_ope = operators == cb_df["GEM_id_source_GEM"].str.split("|").apply(
        match_gem_id_to_operator
    )

    if all(diff_ope):
        return "✅ OK - carbon_bombs_info: Operators are the same than in GEM source\n"
    else:
        return f"❌ KO - carbon_bombs_info: Some operators are different than in GEM source ({operators.loc[~diff_ope].unique()})\n"


def check_carbons_bomb_info(cb_df):
    """Check carbon_bombs_info dataset"""
    txt_res = ""

    txt_res += _check_size(cb_df)

    # merge with CB source to see if every project is available
    merge_df = cb_df.merge(
        cb_source_df,
        left_on=["Project_name", "Country"],
        right_on=["Project Name", "Country"],
        how="inner",
    )

    txt_res += _check_all_cb_proj_available(cb_df, merge_df)
    txt_res += _check_status(cb_df, merge_df)
    txt_res += _check_cb_names(cb_df)
    txt_res += _check_cb_country(merge_df)
    txt_res += _check_cb_emissions(merge_df)
    txt_res += _check_cb_fuel(merge_df)

    # - Project_name
    #     - match all CB name in orig file
    #     - full match cnx cb comp
    # TODO : check for CB with only Others..
    # cb_df.loc[~cb_df["Project_name"].isin(cnx_cb_comp_df["Carbon_bomb_name"])]
    # assert set(cb_df["Project_name"]) == set(cnx_cb_comp_df["Carbon_bomb_name"])

    # Retrieve all units in CB df
    units = cb_df["GEM_id_source_GEM"].str.split("|")
    units = set(itertools.chain(*units))

    txt_res += _check_units_mines_found_in_gem(units)

    txt_res += (
        f"✅ OK - carbon_bombs_info: Nb of 'no information' for GEM_id_source_GEM is "
    )
    txt_res += (
        str(sum(cb_df["GEM_id_source_GEM"] == "No informations available on GEM"))
        + "\n"
    )

    # Retrieve all GEM wiki url in CB df
    url = cb_df["GEM_url_source_GEM"].str.split("|")
    url = set(itertools.chain(*url))

    txt_res += _check_gem_url_and_source_found_in_cb(url, cb_df)
    txt_res += _check_cb_lat_long(cb_df)
    # txt_res += _check_cb_operators(cb_df)

    # - Parent_company_source_GEM :
    #     - show amount of "" or None
    #     - find same companies (TODO)

    # - Companies_involved_source_GEM :
    #     - if Parent_company_source_GEM == "" or None in Parent_company_source_GEM then Parent_company_source_GEM else Operators_source_GEM
    # problem
    # assert all(
    #     cb_df["Companies_involved_source_GEM"] == np.where(
    #         cb_df["Parent_company_source_GEM"] == "No informations on company (100.0%)",
    #         cb_df["Operators_source_GEM"],
    #         cb_df["Parent_company_source_GEM"]
    #     )
    # )

    # - Multiple_unit_concerned_source_GEM :
    #     - not empty only if ; in GEM_id_source_GEM (same number of ; and |)
    #     - if | in Operators_source_GEM not empty (same number of |)
    #     - if | in Parent_company_source_GEM not empty (same number of |)
    # - Status_source_GEM
    #     - match all CB column in orig file
    # - Status_lvl_1
    # - Status_lvl_2

    return txt_res


def check_bank_data(bank_df, cnx_bank_comp_df):
    """Check bank data"""
    banks = set(bank_df["Bank Name"])
    banks_conx = set(cnx_bank_comp_df["Bank"])

    txt = ""

    if len(banks - banks_conx) == 0:
        txt += "✅ OK - bank_data: all banks are in banks to companies connection\n"
    else:
        txt += f"❌ KO - bank_data: some banks are not in banks to companies connection ({banks - banks_conx})\n"

    return txt


def check_connection_bank_company(cnx_bank_comp_df, bank_df, comp_df):
    """Check connection_bank_company"""
    txt = ""

    diff = set(cnx_bank_comp_df["Bank"]) - set(bank_df["Bank Name"])
    if len(diff) == 0:
        txt += "✅ OK - connection_bank_company: all banks in connection bank companies are in banks\n"
    else:
        txt += f"❌ KO - connection_bank_company: some banks in connection bank companies are not in banks ({diff})\n"

    diff = set(cnx_bank_comp_df["Company"]) - set(comp_df["Company_name"])
    if len(diff) == 0:
        txt += "✅ OK - connection_bank_company: all companies in connection bank companies are in companies\n"
    else:
        txt += f"❌ KO - connection_bank_company: some companies in connection bank companies are not in companies ({diff})\n"

    return txt


def check_connection_carbonbombs_company(cnx_cb_comp_df, cb_df, comp_df):
    """Check connection_carbonbombs_company"""
    txt = ""

    diff = set(cnx_cb_comp_df["Carbon_bomb_name"]) - set(cb_df["Project_name"])
    if len(diff) == 0:
        txt += "✅ OK - connection_carbonbombs_company: all CB in connection CB companies are in CB\n"
    else:
        txt += f"❌ KO - connection_carbonbombs_company: some CB in connection CB companies are not in CB ({diff})\n"

    diff = (
        set(cnx_cb_comp_df["Company"].fillna("None"))
        - set(comp_df["Company_name"])
        - {"None", "No informations on company"}
    )
    if len(diff) == 0:
        txt += "✅ OK - connection_carbonbombs_company: all companies in connection CB companies are in companies\n"
    else:
        txt += f"❌ KO - connection_carbonbombs_company: some companies in connection CB companies are not in companies ({diff})\n"

    diff = set(comp_df["Company_name"]) - set(cnx_cb_comp_df["Company"]) - {"None"}
    if len(diff) == 0:
        txt += "✅ OK - connection_carbonbombs_company: all companies in companies are in connection CB companies\n"
    else:
        txt += f"❌ KO - connection_carbonbombs_company: some companies in companies are not in connection CB companies ({diff})\n"

    return txt


def check_metadatas(metadatas, sheet_names):
    """Check metadatas"""
    txt = ""
    for sheet_name in sheet_names:
        if sheet_name == "metadatas":
            continue

        meta_df = metadatas.loc[metadatas["sheetName"] == sheet_name]
        df_ = pd.read_excel(FPATH_OUT_ALL, sheet_name=sheet_name)

        if set(df_.columns) == set(meta_df["columnName"]):
            txt += f"✅ OK - metadatas: For {sheet_name}: columns OK\n"
        else:
            diff = set(df_.columns) - set(meta_df["columnName"])
            txt += f"❌ KO - metadatas: For {sheet_name}: miss some columns ({diff})\n"

    return txt


def check_cleaned_datasets():
    """Apply some checks for all cleaned datasets"""

    f = load_workbook(FPATH_OUT_ALL)
    sheet_names = f.sheetnames

    bank_df = pd.read_excel(FPATH_OUT_ALL, sheet_name="bank_data")
    cb_df = pd.read_excel(FPATH_OUT_ALL, sheet_name="carbon_bombs_data")
    comp_df = pd.read_excel(FPATH_OUT_ALL, sheet_name="company_data")
    cnx_bank_comp_df = pd.read_excel(
        FPATH_OUT_ALL, sheet_name="connection_bank_company"
    )
    cnx_cb_comp_df = pd.read_excel(
        FPATH_OUT_ALL, sheet_name="connection_carbonbombs_company"
    )
    countries_df = pd.read_excel(FPATH_OUT_ALL, sheet_name="country_data")
    metadatas = pd.read_excel(FPATH_OUT_ALL, sheet_name="metadatas")

    res = "\n===== Check cleaned datasets =====\n"
    # deactivate CB check for the moment
    # res += check_carbons_bomb_info(cb_df)
    res += check_bank_data(bank_df, cnx_bank_comp_df)
    res += check_connection_bank_company(cnx_bank_comp_df, bank_df, comp_df)
    res += check_connection_carbonbombs_company(cnx_cb_comp_df, cb_df, comp_df)
    res += check_metadatas(metadatas, sheet_names)
    res += "\n"

    return res
